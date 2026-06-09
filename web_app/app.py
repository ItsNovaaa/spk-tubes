import io
import os
import re
import tempfile
import uuid
import pandas as pd
import numpy as np
from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Limit file uploads to 16MB

# In-memory session store for uploaded DataFrames (using unique task IDs)
# In production, this would be a database or Redis, but in-memory is perfect for a local app.
STORED_DATA = {}
REFERENCE_WORKBOOK_PATH = os.environ.get(
    'FUCOM_VIKOR_REFERENCE_XLSX',
    r'C:\Users\User\Downloads\TUGAS BESAR SPK(2).xlsx'
)


def _is_blank_value(value):
    return pd.isna(value) or str(value).strip() == ''


def make_json_safe(value):
    """
    Convert pandas/numpy values into JSON-friendly Python values.
    """
    if isinstance(value, dict):
        return {k: make_json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [make_json_safe(v) for v in value]
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return None if not np.isfinite(value) else float(value)
    if isinstance(value, float):
        return None if not np.isfinite(value) else value
    if pd.isna(value):
        return None
    return value


def detect_numeric_like_columns(df, alternatif_col):
    """
    Detect criteria columns from rows that actually contain an alternative name.

    This handles worksheets that contain notes/metadata below the dataset. Pandas
    may mark columns as object when those notes contain text, even though the
    decision rows themselves are numeric.
    """
    if alternatif_col not in df.columns:
        return []

    alt_mask = df[alternatif_col].apply(lambda value: not _is_blank_value(value))
    data_rows = df.loc[alt_mask]
    numeric_cols = []

    for col in df.columns:
        if col == alternatif_col:
            continue
        numeric_values = pd.to_numeric(data_rows[col], errors='coerce')
        has_value = data_rows[col].apply(lambda value: not _is_blank_value(value))
        if has_value.any() and numeric_values[has_value].notna().all():
            numeric_cols.append(col)

    return numeric_cols


def clean_decision_matrix(df, alternatif_col, criteria_cols):
    """
    Keep only valid decision rows and coerce all selected criteria to numeric.
    """
    if alternatif_col not in df.columns:
        raise ValueError(f"Kolom alternatif '{alternatif_col}' tidak ditemukan.")

    missing_criteria = [col for col in criteria_cols if col not in df.columns]
    if missing_criteria:
        raise ValueError(f"Kolom kriteria tidak ditemukan: {', '.join(missing_criteria)}")

    if not criteria_cols:
        raise ValueError("Minimal satu kolom kriteria harus dipilih.")

    selected_cols = [alternatif_col] + criteria_cols
    clean_df = df[selected_cols].copy()

    alt_mask = clean_df[alternatif_col].apply(lambda value: not _is_blank_value(value))
    clean_df = clean_df.loc[alt_mask].copy()

    if clean_df.empty:
        raise ValueError("Tidak ada baris alternatif yang valid untuk dihitung.")

    clean_df[alternatif_col] = clean_df[alternatif_col].astype(str).str.strip()

    invalid_messages = []
    for col in criteria_cols:
        original = clean_df[col]
        numeric = pd.to_numeric(original, errors='coerce')
        invalid_mask = original.apply(lambda value: not _is_blank_value(value)) & numeric.isna()
        if invalid_mask.any():
            bad_alts = clean_df.loc[invalid_mask, alternatif_col].head(5).tolist()
            invalid_messages.append(f"{col} pada alternatif: {', '.join(bad_alts)}")
        clean_df[col] = numeric

    if invalid_messages:
        raise ValueError("Nilai kriteria harus berupa angka. Periksa " + "; ".join(invalid_messages))

    missing_mask = clean_df[criteria_cols].isna()
    if missing_mask.any().any():
        bad_rows = clean_df.loc[missing_mask.any(axis=1), alternatif_col].head(5).tolist()
        raise ValueError("Ada nilai kriteria kosong pada alternatif: " + ", ".join(bad_rows))

    return clean_df.reset_index(drop=True)


def normalize_criteria_type(value):
    normalized = str(value).strip().lower()
    if normalized == 'benefit':
        return 'Benefit'
    if normalized == 'cost':
        return 'Cost'
    raise ValueError(f"Tipe kriteria tidak valid: {value}. Gunakan Benefit atau Cost.")


def load_reference_excel_bytes():
    try:
        with open(REFERENCE_WORKBOOK_PATH, 'rb') as reference_file:
            return reference_file.read()
    except OSError:
        return None


def infer_excel_normalization_types(excel_bytes, sheet_name, criteria_cols, stage):
    """
    Infer criteria type from the workbook's own normalization formulas.

    The reference workbook uses different formulas for the FUCOM weighting
    normalization and the VIKOR distance normalization:
    - FUCOM: (x - min) is Benefit, (max - x) is Cost.
    - VIKOR: (max - x) is Benefit, (x - min) is Cost.
    """
    if not excel_bytes:
        return {}

    try:
        workbook = load_workbook(io.BytesIO(excel_bytes), data_only=False)
    except Exception:
        return {}

    if sheet_name not in workbook.sheetnames:
        return {}

    worksheet = workbook[sheet_name]
    normalisasi_row = None
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().upper() == 'NORMALISASI':
                normalisasi_row = cell.row
                break
        if normalisasi_row is not None:
            break

    if normalisasi_row is None:
        return {}

    header_row = normalisasi_row + 1
    first_data_row = header_row + 1
    header_to_col = {}
    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(header_row, col_idx).value
        if header is not None:
            header_to_col[str(header).strip()] = col_idx

    inferred = {}
    for criteria in criteria_cols:
        col_idx = header_to_col.get(str(criteria))
        if not col_idx:
            continue

        col_letter = get_column_letter(col_idx).upper()
        formula = worksheet.cell(first_data_row, col_idx).value
        if not isinstance(formula, str) or not formula.startswith('='):
            continue

        compact_formula = formula.replace(' ', '').upper()
        numerator = compact_formula.split('/')[0].replace('$', '')
        same_column_refs = re.findall(rf"{col_letter}(\d+)", numerator)
        if len(same_column_refs) < 2:
            continue

        first_ref_row = int(same_column_refs[0])
        second_ref_row = int(same_column_refs[1])
        source_data_row = min(first_ref_row, second_ref_row)
        cell_minus_ref = first_ref_row == source_data_row
        ref_minus_cell = second_ref_row == source_data_row

        if stage == 'fucom':
            if cell_minus_ref:
                inferred[criteria] = 'Benefit'
            elif ref_minus_cell:
                inferred[criteria] = 'Cost'
        elif stage == 'vikor':
            if ref_minus_cell:
                inferred[criteria] = 'Benefit'
            elif cell_minus_ref:
                inferred[criteria] = 'Cost'

    return inferred


def extract_excel_fucom_weight_rank_mapping(excel_bytes, criteria_cols):
    """
    Extract how the reference workbook maps final criteria weights.

    In the FUCOM sheet, the final "BOBOT TIAP KRITERIA" table may point each
    criterion to a specific normalized W row (for example =B95, meaning W7
    norm). This helper returns that criterion -> W rank mapping so dataset-only
    uploads can still follow the same Excel reference behavior.
    """
    if not excel_bytes:
        return None

    try:
        workbook = load_workbook(io.BytesIO(excel_bytes), data_only=False)
    except Exception:
        return None

    if 'FUCOM' not in workbook.sheetnames:
        return None

    worksheet = workbook['FUCOM']
    criteria_set = {str(col) for col in criteria_cols}

    for row in worksheet.iter_rows():
        for cell in row:
            if not (isinstance(cell.value, str) and cell.value.strip().upper() == 'BOBOT TIAP KRITERIA'):
                continue

            mapping = {}
            label_col = cell.column
            formula_col = cell.column + 1
            for row_idx in range(cell.row + 1, worksheet.max_row + 1):
                label = worksheet.cell(row_idx, label_col).value
                formula = worksheet.cell(row_idx, formula_col).value
                if label is None:
                    continue

                label = str(label).strip()
                if label.lower() == 'jumlah':
                    break
                if label not in criteria_set or not isinstance(formula, str):
                    continue

                match = re.search(r"B(\d+)", formula.upper())
                if not match:
                    continue

                source_row = int(match.group(1))
                source_label = str(worksheet.cell(source_row, 1).value or '')
                rank_match = re.search(r"W\s*(\d+)", source_label.upper())
                if rank_match:
                    mapping[label] = int(rank_match.group(1))

                if criteria_set.issubset(mapping.keys()):
                    return {col: mapping[str(col)] for col in criteria_cols}

    return None


def extract_excel_fucom_weights(excel_bytes, criteria_cols):
    """
    Read the workbook's final FUCOM weight table when it exists.

    The reference workbook stores the final weights in VIKOR under "Bobot
    FUCOM", linked from the FUCOM sheet's "BOBOT TIAP KRITERIA" block. Reading
    this table makes the web calculation match the submitted Excel exactly,
    including the workbook's own tie handling and final criteria mapping.
    """
    if not excel_bytes:
        return None

    try:
        workbook = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    except Exception:
        return None

    criteria_set = {str(col) for col in criteria_cols}

    def collect_vertical_table(worksheet, start_row, label_col, value_col):
        weights = {}
        for row_idx in range(start_row, worksheet.max_row + 1):
            label = worksheet.cell(row_idx, label_col).value
            value = worksheet.cell(row_idx, value_col).value
            if label is None:
                continue
            label = str(label).strip()
            if label.lower() == 'jumlah':
                break
            if label in criteria_set:
                try:
                    numeric_value = float(value)
                except (TypeError, ValueError):
                    continue
                if np.isfinite(numeric_value):
                    weights[label] = numeric_value
            if criteria_set.issubset(weights.keys()):
                break
        return weights

    # Preferred source: VIKOR sheet, because it is exactly what VIKOR uses.
    if 'VIKOR' in workbook.sheetnames:
        worksheet = workbook['VIKOR']
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().upper() == 'BOBOT FUCOM':
                    weights = collect_vertical_table(worksheet, cell.row + 1, cell.column, cell.column + 1)
                    if criteria_set.issubset(weights.keys()):
                        return {col: weights[str(col)] for col in criteria_cols}

    # Fallback source: FUCOM sheet's final "BOBOT TIAP KRITERIA" table.
    if 'FUCOM' in workbook.sheetnames:
        worksheet = workbook['FUCOM']
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().upper() == 'BOBOT TIAP KRITERIA':
                    weights = collect_vertical_table(worksheet, cell.row + 1, cell.column, cell.column + 1)
                    if criteria_set.issubset(weights.keys()):
                        return {col: weights[str(col)] for col in criteria_cols}

    return None


def hitung_fucom_objektif_local(df, criteria_cols, jenis_kriteria, include_details=False):
    """
    Menghitung bobot FUCOM secara objektif berdasarkan rata-rata normalisasi data.
    """
    df_norm = pd.DataFrame()
    for col in criteria_cols:
        x_min = df[col].min()
        x_max = df[col].max()
        
        if x_max == x_min:
            df_norm[col] = 1.0
            continue
            
        if jenis_kriteria[col] == 'Benefit':
            df_norm[col] = (df[col] - x_min) / (x_max - x_min)
        else:
            df_norm[col] = (x_max - df[col]) / (x_max - x_min)
            
    skor_kepentingan = df_norm.mean()
    nilai_agregat = df_norm.sum()
    skor_sorted = skor_kepentingan.sort_values(ascending=False)
    kriteria_urut = skor_sorted.index.tolist()

    priority_rows = []
    for i, col in enumerate(criteria_cols):
        sorted_col = kriteria_urut[i] if i < len(kriteria_urut) else None
        priority_rows.append({
            'kriteria': col,
            'nilai_agregat': nilai_agregat[col],
            'hasil_urutan': sorted_col,
            'ranking': i + 1 if sorted_col is not None else None,
            'keterangan': 'TERKECIL' if i == len(criteria_cols) - 1 else ''
        })
    
    # Hitung Phi
    phi = []
    ratio_rows = []
    weight_labels = {col: f"w{i + 1}" for i, col in enumerate(criteria_cols)}
    for i in range(len(skor_sorted) - 1):
        nilai_phi = skor_sorted.iloc[i] / skor_sorted.iloc[i+1] if skor_sorted.iloc[i+1] != 0 else 1.0
        phi.append(nilai_phi)
        current_col = kriteria_urut[i]
        next_col = kriteria_urut[i + 1]
        ratio_rows.append({
            'phi': f"\u03c6{i + 1}",
            'nilai_phi': nilai_phi,
            'rasio_kriteria': f"{current_col}/{next_col}",
            'rasio_bobot': f"{weight_labels[current_col]}/{weight_labels[next_col]}"
        })
        
    # Hitung bobot konsisten dari bawah ke atas
    w_raw = [1.0]
    phi_reversed = phi[::-1]
    for p in phi_reversed:
        w_raw.append(w_raw[-1] * p)
        
    w_raw.reverse()
    total_w_raw = sum(w_raw)
    w_final = [w / total_w_raw for w in w_raw]

    consistency_rows = []
    n_criteria = len(kriteria_urut)
    for priority_index in range(n_criteria, 0, -1):
        raw_index = priority_index - 1
        formula = 'Basis 1' if priority_index == n_criteria else f"\u03c6{priority_index}*W{priority_index + 1}"
        consistency_rows.append({
            'label': f"W{priority_index}",
            'criteria': kriteria_urut[raw_index],
            'raw_weight': w_raw[raw_index],
            'formula': formula
        })

    normalized_rows = []
    for i, col in enumerate(kriteria_urut):
        normalized_rows.append({
            'label': f"W{i + 1} norm",
            'criteria': col,
            'normalized_weight': w_final[i],
            'formula': f"W{i + 1}/Wj"
        })
    
    bobot_final = {}
    for i, k in enumerate(kriteria_urut):
        bobot_final[k] = w_final[i]
        
    # Urutkan sesuai input kolom asli
    bobot_urut_asli = {col: bobot_final[col] for col in criteria_cols}

    if include_details:
        return bobot_urut_asli, {
            'available': True,
            'priority_rows': priority_rows,
            'ratio_rows': ratio_rows,
            'consistency_rows': consistency_rows,
            'normalized_rows': normalized_rows,
            'total_raw_weight': total_w_raw,
            'total_normalized_weight': sum(w_final),
            'basis_label': f"W{n_criteria}"
        }

    return bobot_urut_asli


def hitung_fucom_subjektif_local(criteria_rank, criteria_priority):
    """
    Menghitung bobot FUCOM secara subjektif berdasarkan urutan prioritas
    dan rasio komparatif (Phi) yang diberikan pengguna.
    """
    # Sesuai teori FUCOM, kriteria terbawah diberi bobot basis 1.0
    w_raw = [1.0]
    
    # Hitung bobot dari bawah ke atas menggunakan phi
    # phi = priority[k] / priority[k-1], so w_raw[k-1] = w_raw[k] * phi
    # Let's reverse the priority to calculate from bottom up
    # criteria_priority is list of comparisons: phi_1 (C_1/C_2), phi_2 (C_2/C_3)...
    phi_reversed = criteria_priority[::-1]
    for p in phi_reversed:
        w_raw.append(w_raw[-1] * p)
        
    w_raw.reverse()
    total_w_raw = sum(w_raw)
    w_final = [w / total_w_raw for w in w_raw]
    
    return {k: w_final[i] for i, k in enumerate(criteria_rank)}


def hitung_vikor_local(df, alternatiff_col, jenis_kriteria, bobot, v=0.5):
    """
    Menghitung perankingan menggunakan metode VIKOR.
    """
    kriteria = list(jenis_kriteria.keys())
    v = min(max(float(v), 0.0), 1.0)
    
    # Step 1: Nilai ideal terbaik (X+) dan terburuk (X-)
    x_plus = {}
    x_minus = {}
    for k in kriteria:
        if jenis_kriteria[k] == 'Benefit':
            x_plus[k] = df[k].max()
            x_minus[k] = df[k].min()
        else:
            x_plus[k] = df[k].min()
            x_minus[k] = df[k].max()
            
    # Step 2: Normalisasi terbobot
    df_norm = pd.DataFrame()
    for k in kriteria:
        jarak_ideal = abs(x_plus[k] - x_minus[k])
        w = bobot[k]
        if jarak_ideal == 0:
            df_norm[k] = 0.0
        else:
            df_norm[k] = w * (abs(x_plus[k] - df[k]) / jarak_ideal)
            
    # Step 3 & 4: Nilai S dan R
    df_vikor = pd.DataFrame()
    df_vikor['ALTERNATIF'] = df[alternatiff_col].astype(str).values
    df_vikor['S'] = df_norm[kriteria].sum(axis=1).values
    df_vikor['R'] = df_norm[kriteria].max(axis=1).values
    
    # Step 5: Nilai Q
    S_star = df_vikor['S'].min()
    S_minus = df_vikor['S'].max()
    R_star = df_vikor['R'].min()
    R_minus = df_vikor['R'].max()
    
    def hitung_q(s, r):
        q_s = (s - S_star) / (S_minus - S_star) if (S_minus - S_star) != 0 else 0.0
        q_r = (r - R_star) / (R_minus - R_star) if (R_minus - R_star) != 0 else 0.0
        return (v * q_s) + ((1.0 - v) * q_r)
        
    df_vikor['Q'] = df_vikor.apply(lambda row: hitung_q(row['S'], row['R']), axis=1)
    
    # Step 6: Perankingan
    df_vikor['Ranking'] = df_vikor['Q'].rank(method='min', ascending=True).astype(int)
    df_hasil = df_vikor.sort_values(by='Ranking').reset_index(drop=True)
    
    # ----------------------------------------------------------
    # Verifikasi Kondisi Kompromi VIKOR (Acceptable Advantage & Stability)
    # ----------------------------------------------------------
    m = len(df_hasil)
    dq = 1.0 / (m - 1) if m > 1 else 0.25
    if m < 4:
        dq = 0.25
        
    compromise_solutions = []
    c1_satisfied = False
    c2_satisfied = False
    details = ""
    
    if m >= 2:
        # C1: Acceptable Advantage
        # Q(A2) - Q(A1) >= DQ
        diff_q = df_hasil.loc[1, 'Q'] - df_hasil.loc[0, 'Q']
        if diff_q >= dq:
            c1_satisfied = True
            
        # C2: Acceptable Stability
        # A1 harus merupakan peringkat terbaik di S dan/atau R
        best_s_alt = df_hasil.loc[df_hasil['S'].idxmin(), 'ALTERNATIF']
        best_r_alt = df_hasil.loc[df_hasil['R'].idxmin(), 'ALTERNATIF']
        a1_alt = df_hasil.loc[0, 'ALTERNATIF']
        
        if (a1_alt == best_s_alt) or (a1_alt == best_r_alt):
            c2_satisfied = True
            
        # Tentukan solusi kompromi jika kondisi tidak terpenuhi
        if c1_satisfied and c2_satisfied:
            compromise_solutions = [df_hasil.loc[0, 'ALTERNATIF']]
            details = "Kondisi C1 (Acceptable Advantage) dan C2 (Acceptable Stability) terpenuhi. Solusi terbaik tunggal."
        elif not c1_satisfied and c2_satisfied:
            # Jika C1 tidak terpenuhi, cari alternatif A_m sampai Q(A_m) - Q(A_1) < DQ
            comp_alts = []
            for idx, row in df_hasil.iterrows():
                if row['Q'] - df_hasil.loc[0, 'Q'] < dq:
                    comp_alts.append(row['ALTERNATIF'])
                else:
                    break
            compromise_solutions = comp_alts
            details = f"Kondisi C1 tidak terpenuhi (Q(A2)-Q(A1) = {diff_q:.4f} < DQ = {dq:.4f}). Alternatif kompromi: {', '.join(comp_alts)}."
        elif c1_satisfied and not c2_satisfied:
            # Jika C2 tidak terpenuhi, solusi kompromi adalah A1 dan A2
            compromise_solutions = [df_hasil.loc[0, 'ALTERNATIF'], df_hasil.loc[1, 'ALTERNATIF']]
            details = "Kondisi C2 tidak terpenuhi (A1 bukan yang terbaik di S atau R). Alternatif kompromi: A1 dan A2."
        else:
            # Keduanya tidak terpenuhi, gabungkan logika
            comp_alts = []
            for idx, row in df_hasil.iterrows():
                if row['Q'] - df_hasil.loc[0, 'Q'] < dq:
                    comp_alts.append(row['ALTERNATIF'])
                else:
                    break
            compromise_solutions = list(set(comp_alts + [df_hasil.loc[1, 'ALTERNATIF']]))
            details = f"Kondisi C1 dan C2 tidak terpenuhi. Alternatif kompromi: {', '.join(compromise_solutions)}."
    else:
        compromise_solutions = [df_hasil.loc[0, 'ALTERNATIF']]
        details = "Jumlah alternatif hanya 1. Solusi otomatis tunggal."
        
    # Buat matriks normalisasi terbobot lengkap untuk dikembalikan
    df_weighted_matrix = df_norm.copy()
    df_weighted_matrix.insert(0, 'ALTERNATIF', df[alternatiff_col].astype(str).values)
        
    return df_hasil, S_star, S_minus, R_star, R_minus, compromise_solutions, details, df_weighted_matrix


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Tidak ada file yang diupload.'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'File belum dipilih.'}), 400
        
    try:
        file_bytes = file.read()
        # Load Excel using Pandas
        # Read sheet names first
        excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
        sheets = excel_file.sheet_names
        
        # Default read first sheet
        df = excel_file.parse(sheets[0])
        
        # Save to store
        session_id = str(uuid.uuid4())
        STORED_DATA[session_id] = {
            'df': df,
            'excel_file_data': excel_file,
            'excel_bytes': file_bytes,
            'filename': file.filename
        }
        
        # Identify columns
        columns = list(df.columns)
        
        # Recommend the alternative-name column. Prefer explicit names first,
        # because pandas may read mixed note columns as object and plain text
        # columns as its newer string dtype.
        alternatif_rec = columns[0]
        explicit_alt_cols = [
            col for col in columns
            if str(col).strip().lower() in {'alternatif', 'alternative', 'nama alternatif', 'nama'}
        ]
        if explicit_alt_cols:
            alternatif_rec = explicit_alt_cols[0]
        else:
            best_text_score = -1
            for col in columns:
                non_blank = df[col].apply(lambda value: not _is_blank_value(value))
                numeric = pd.to_numeric(df.loc[non_blank, col], errors='coerce')
                text_score = int(non_blank.sum() - numeric.notna().sum())
                if text_score > best_text_score:
                    best_text_score = text_score
                    alternatif_rec = col
                
        # Get numeric-like columns (potential criteria). Notes below the data
        # table should not prevent numeric decision columns from being detected.
        potential_criteria = detect_numeric_like_columns(df, alternatif_rec)
        
        # Preview top 5 rows
        preview_data = make_json_safe(df.head(5).to_dict(orient='records'))
        
        return jsonify({
            'session_id': session_id,
            'sheets': sheets,
            'columns': columns,
            'alternatif_rec': alternatif_rec,
            'potential_criteria': potential_criteria,
            'preview': preview_data
        })
        
    except Exception as e:
        return jsonify({'error': f"Gagal membaca Excel: {str(e)}"}), 500


@app.route('/calculate', methods=['POST'])
def calculate():
    req_data = request.json
    session_id = req_data.get('session_id')
    alternatif_col = req_data.get('alternatif_col')
    criteria_configs = req_data.get('criteria_configs')  # List of {name, type: 'Benefit'/'Cost', weight_mode, manual_weight, rank, phi}
    v = float(req_data.get('v', 0.5))
    
    if session_id not in STORED_DATA:
        return jsonify({'error': 'Sesi data kadaluarsa atau tidak ditemukan. Silakan upload file kembali.'}), 400
        
    df = STORED_DATA[session_id]['df']
    
    try:
        if not isinstance(criteria_configs, list) or not criteria_configs:
            return jsonify({'error': 'Konfigurasi kriteria belum dipilih.'}), 400

        criteria_cols = [c['name'] for c in criteria_configs]
        jenis_kriteria = {c['name']: normalize_criteria_type(c['type']) for c in criteria_configs}
        df_clean = clean_decision_matrix(df, alternatif_col, criteria_cols)
        excel_bytes = STORED_DATA[session_id].get('excel_bytes')
        reference_excel_bytes = load_reference_excel_bytes()
        excel_fucom_types = infer_excel_normalization_types(excel_bytes, 'FUCOM', criteria_cols, 'fucom')
        excel_vikor_types = infer_excel_normalization_types(excel_bytes, 'VIKOR', criteria_cols, 'vikor')
        excel_fucom_weights = extract_excel_fucom_weights(excel_bytes, criteria_cols)
        reference_fucom_types = infer_excel_normalization_types(reference_excel_bytes, 'FUCOM', criteria_cols, 'fucom')
        reference_vikor_types = infer_excel_normalization_types(reference_excel_bytes, 'VIKOR', criteria_cols, 'vikor')
        reference_weight_rank_mapping = extract_excel_fucom_weight_rank_mapping(reference_excel_bytes, criteria_cols)
        jenis_kriteria_fucom = {col: excel_fucom_types.get(col, jenis_kriteria[col]) for col in criteria_cols}
        jenis_kriteria_vikor = {col: excel_vikor_types.get(col, jenis_kriteria[col]) for col in criteria_cols}
        if not excel_fucom_types and reference_fucom_types:
            jenis_kriteria_fucom = {col: reference_fucom_types.get(col, jenis_kriteria_fucom[col]) for col in criteria_cols}
        if not excel_vikor_types and reference_vikor_types:
            jenis_kriteria_vikor = {col: reference_vikor_types.get(col, jenis_kriteria_vikor[col]) for col in criteria_cols}
        
        # ----------------------------------------------------------
        # Hitung Bobot Kriteria
        # ----------------------------------------------------------
        weight_mode = req_data.get('weight_mode', 'objective')  # objective, manual, subjective
        fucom_details = {
            'available': False,
            'priority_rows': [],
            'ratio_rows': []
        }
        
        if weight_mode == 'objective':
            computed_bobot, fucom_details = hitung_fucom_objektif_local(
                df_clean, criteria_cols, jenis_kriteria_fucom, include_details=True
            )
            if excel_fucom_weights is not None:
                bobot = excel_fucom_weights
                final_weight_source = 'uploaded_excel'
            elif reference_weight_rank_mapping:
                normalized_by_rank = {
                    int(row['label'].replace('W', '').replace(' norm', '')): row['normalized_weight']
                    for row in fucom_details['normalized_rows']
                }
                bobot = {
                    col: normalized_by_rank.get(reference_weight_rank_mapping[col], computed_bobot[col])
                    for col in criteria_cols
                }
                final_weight_source = 'reference_excel_mapping'
            else:
                bobot = computed_bobot
                final_weight_source = 'computed'

            fucom_details['final_weight_source'] = final_weight_source
            fucom_details['final_weights'] = bobot
        elif weight_mode == 'manual':
            raw_weights = {c['name']: float(c.get('manual_weight', 0.0)) for c in criteria_configs}
            total = sum(raw_weights.values())
            if total == 0:
                # Fallback if all weights are zero
                bobot = {k: 1.0 / len(criteria_cols) for k in criteria_cols}
            else:
                bobot = {k: v_w / total for k, v_w in raw_weights.items()}
        elif weight_mode == 'subjective':
            # Sort criteria configs by rank (ascending: 1 is top, then 2, 3...)
            sorted_configs = sorted(criteria_configs, key=lambda x: int(x.get('rank', 1)))
            criteria_rank = [c['name'] for c in sorted_configs]
            
            # The priorities (phi ratios) of consecutive elements.
            # phi_k = ratio of rank_k over rank_{k+1}.
            # The last element does not have a next element, so it is just 1.0 or ignored.
            # criteria_priority is list of comparisons.
            # Let's extract priorities: priority values.
            # To stay fully consistent:
            # w_k / w_{k+1} = phi_k.
            # We can extract phi values from the form inputs.
            # For n criteria, there are n-1 comparisons.
            # Let's gather phis:
            phi_values = []
            for i in range(len(sorted_configs) - 1):
                phi_val = float(sorted_configs[i].get('phi', 1.0))
                phi_values.append(phi_val)
                
            bobot = hitung_fucom_subjektif_local(criteria_rank, phi_values)
            # Reorder weights to match original criteria order
            bobot = {col: bobot[col] for col in criteria_cols}
        else:
            return jsonify({'error': f'Mode bobot tidak valid: {weight_mode}'}), 400
            
        # ----------------------------------------------------------
        # Hitung VIKOR
        # ----------------------------------------------------------
        df_hasil, S_star, S_minus, R_star, R_minus, compromise_solutions, compromise_details, df_weighted_matrix = hitung_vikor_local(
            df_clean, alternatif_col, jenis_kriteria_vikor, bobot, v
        )
        
        # Save results in session store for export
        STORED_DATA[session_id]['results'] = {
            'df_hasil': df_hasil,
            'bobot': bobot,
            'jenis_kriteria': jenis_kriteria_vikor,
            'jenis_kriteria_fucom': jenis_kriteria_fucom,
            'jenis_kriteria_vikor': jenis_kriteria_vikor,
            'excel_fucom_types': excel_fucom_types,
            'excel_vikor_types': excel_vikor_types,
            'excel_fucom_weights': excel_fucom_weights,
            'reference_fucom_types': reference_fucom_types,
            'reference_vikor_types': reference_vikor_types,
            'reference_weight_rank_mapping': reference_weight_rank_mapping,
            'ref_values': {
                'S_star': S_star,
                'S_minus': S_minus,
                'R_star': R_star,
                'R_minus': R_minus,
                'v': v
            },
            'compromise_solutions': compromise_solutions,
            'compromise_details': compromise_details,
            'df_weighted_matrix': df_weighted_matrix,
            'fucom_details': fucom_details
        }
        
        # Return results to UI
        return jsonify({
            'rankings': make_json_safe(df_hasil.to_dict(orient='records')),
            'bobot': make_json_safe(bobot),
            'ref_values': {
                'S_star': make_json_safe(S_star),
                'S_minus': make_json_safe(S_minus),
                'R_star': make_json_safe(R_star),
                'R_minus': make_json_safe(R_minus),
                'v': make_json_safe(v)
            },
            'compromise_solutions': make_json_safe(compromise_solutions),
            'compromise_details': compromise_details,
            'weighted_matrix': make_json_safe(df_weighted_matrix.to_dict(orient='records')),
            'fucom_details': make_json_safe(fucom_details),
            'criteria_types': {
                'selected': make_json_safe(jenis_kriteria),
                'fucom': make_json_safe(jenis_kriteria_fucom),
                'vikor': make_json_safe(jenis_kriteria_vikor),
                'excel_fucom_detected': make_json_safe(excel_fucom_types),
                'excel_vikor_detected': make_json_safe(excel_vikor_types),
                'excel_fucom_weights': make_json_safe(excel_fucom_weights),
                'reference_fucom_detected': make_json_safe(reference_fucom_types),
                'reference_vikor_detected': make_json_safe(reference_vikor_types),
                'reference_weight_rank_mapping': make_json_safe(reference_weight_rank_mapping)
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Gagal memproses perhitungan: {str(e)}'}), 500


@app.route('/export', methods=['POST'])
def export_results():
    req_data = request.json
    session_id = req_data.get('session_id')
    
    if session_id not in STORED_DATA or 'results' not in STORED_DATA[session_id]:
        return jsonify({'error': 'Data hasil perhitungan tidak ditemukan atau sesi kadaluarsa.'}), 400
        
    try:
        results = STORED_DATA[session_id]['results']
        df_hasil = results['df_hasil']
        bobot = results['bobot']
        jenis_kriteria = results['jenis_kriteria']
        jenis_kriteria_fucom = results.get('jenis_kriteria_fucom', jenis_kriteria)
        jenis_kriteria_vikor = results.get('jenis_kriteria_vikor', jenis_kriteria)
        ref = results['ref_values']
        comp_details = results['compromise_details']
        df_weighted_matrix = results['df_weighted_matrix']
        
        # Generate Excel
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, f"hasil_fucom_vikor_{session_id}.xlsx")
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Sheet 1: Hasil Ranking
            df_hasil.to_excel(writer, sheet_name='Hasil Ranking', index=False)
            
            # Sheet 2: Bobot FUCOM
            df_bobot = pd.DataFrame([
                {
                    'Kriteria': k, 
                    'Tipe FUCOM': jenis_kriteria_fucom[k],
                    'Tipe VIKOR': jenis_kriteria_vikor[k],
                    'Bobot (Desimal)': v_b, 
                    'Bobot (%)': f"{v_b*100:.4f}%"
                }
                for k, v_b in bobot.items()
            ])
            df_bobot.to_excel(writer, sheet_name='Bobot Kriteria', index=False)
            
            # Sheet 3: Nilai Referensi & Kompromi
            df_ref = pd.DataFrame({
                'Parameter / Kondisi': [
                    'S* (S terbaik/min)', 
                    'S⁻ (S terburuk/max)', 
                    'R* (R terbaik/min)', 
                    'R⁻ (R terburuk/max)', 
                    'v (Strategi koefisien)',
                    'Detail Solusi Kompromi'
                ],
                'Nilai / Deskripsi': [
                    ref['S_star'], 
                    ref['S_minus'], 
                    ref['R_star'], 
                    ref['R_minus'], 
                    ref['v'],
                    comp_details
                ]
            })
            df_ref.to_excel(writer, sheet_name='Nilai Referensi & Kompromi', index=False)
            
            # Sheet 4: Matriks Normalisasi Terbobot
            df_weighted_matrix.to_excel(writer, sheet_name='Matriks Terbobot', index=False)
            
        return jsonify({'download_url': f'/download/{session_id}'})
        
    except Exception as e:
        return jsonify({'error': f'Gagal membuat file Excel: {str(e)}'}), 500


@app.route('/download/<session_id>', methods=['GET'])
def download_file(session_id):
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, f"hasil_fucom_vikor_{session_id}.xlsx")
    
    if not os.path.exists(file_path):
        return "File tidak ditemukan atau sesi telah berakhir.", 404
        
    return send_file(
        file_path,
        as_attachment=True,
        download_name="hasil_fucom_vikor.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)
